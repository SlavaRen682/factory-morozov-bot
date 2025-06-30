import telebot
from telebot import types
import os
import flask
import datetime
from openpyxl import Workbook, load_workbook

TOKEN = os.environ.get("TOKEN")
if not TOKEN:
    raise ValueError("Ошибка: переменная окружения TOKEN не установлена!")

owner_id_str = os.environ.get("OWNER_ID")
if not owner_id_str:
    raise ValueError("Ошибка: переменная окружения OWNER_ID не установлена!")
OWNER_ID = int(owner_id_str)

bot = telebot.TeleBot(TOKEN, threaded=False)
STATE = {}
DATA = {}
PHOTO_LINK = {}
EXCEL_FILE = 'orders.xlsx'

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Имя", "Username", "Фото", "Реквизиты"])
        wb.save(EXCEL_FILE)

def save_to_excel(user, photo_path, requisites):
    ensure_excel_file()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        user.first_name,
        f"@{user.username}" if user.username else "—",
        photo_path,
        requisites
    ])
    wb.save(EXCEL_FILE)

@bot.message_handler(commands=['start'])
def start(message):
    name = message.from_user.first_name or "друг"

    welcome_text = (
        f"👋 Привет, <b>{name}</b>!\n\n"
        f"Вы попали в официальный бот фабрики <b>Морозовых</b> 🧵\n"
        f"Здесь вы сможете:\n"
        f"▫️ Отправить фото изделия\n"
        f"▫️ Получить примерную стоимость пошива\n"
        f"▫️ Передать реквизиты для оформления\n\n"
        f"<i>Актуален ли для вас индивидуальный пошив изделий?</i>"
    )

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add("Да", "Нет")

    bot.send_message(
        message.chat.id,
        welcome_text,
        reply_markup=markup,
        parse_mode='HTML'
    )

    STATE[message.chat.id] = 'AWAIT_CONFIRM'


@bot.message_handler(commands=['contact'])
def contact_command(message):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("📲 Перейти в Telegram", url="https://t.me/xatyba"))
    bot.send_message(message.chat.id, "💬 Менеджер на связи по кнопке ниже:", reply_markup=markup)

@bot.message_handler(commands=['excel', 'клиенты'])
def send_excel_to_owner(message):
    if message.chat.id == OWNER_ID and os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, 'rb') as f:
            bot.send_document(OWNER_ID, f, caption="📊 Актуальный Excel-файл с заявками")
    else:
        bot.send_message(message.chat.id, "Файл не найден или у вас нет прав.")

@bot.message_handler(func=lambda m: STATE.get(m.chat.id) == 'AWAIT_CONFIRM')
def confirm_interest(message):
    if message.text.lower() == "да":
        bot.send_message(message.chat.id, "📸 Пришлите фото изделия, и мы передадим его на оценку.", reply_markup=types.ReplyKeyboardRemove())
        STATE[message.chat.id] = 'AWAIT_PHOTO'
    else:
        bot.send_message(message.chat.id, "Хорошо! Если заинтересует — возвращайтесь 🙂", reply_markup=types.ReplyKeyboardRemove())
        STATE.pop(message.chat.id, None)

@bot.message_handler(content_types=['photo'])
def handle_photo(message):
    if STATE.get(message.chat.id) != 'AWAIT_PHOTO':
        return

    user = message.from_user
    file_info = bot.get_file(message.photo[-1].file_id)
    downloaded = bot.download_file(file_info.file_path)

    os.makedirs("photos", exist_ok=True)
    photo_name = f"photo_{user.id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
    photo_path = os.path.join("photos", photo_name)
    with open(photo_path, 'wb') as f:
        f.write(downloaded)

    DATA[message.chat.id] = {"photo_path": photo_path, "user": user}
    STATE[message.chat.id] = 'WAITING_OWNER_PRICE'

    bot.send_message(message.chat.id, "✅ Фото получено. Мы передали его на оценку, ожидайте ответ с примерной ценой.")

    caption = (
        f"🆕 Фото от клиента:\n"
        f"👤 {user.first_name} (@{user.username or '—'})\n"
        f"🆔 ID: {user.id}\n\n"
        f"✍ Ответьте на это сообщение примерной ценой."
    )
    try:
        with open(photo_path, 'rb') as photo:
            sent = bot.send_photo(OWNER_ID, photo, caption=caption)
            PHOTO_LINK[sent.message_id] = message.chat.id
    except Exception as e:
        print(f"❌ Ошибка отправки фото владельцу: {e}")

@bot.message_handler(func=lambda m: m.chat.id == OWNER_ID and m.reply_to_message)
def handle_owner_reply(message):
    reply_id = message.reply_to_message.message_id
    client_id = PHOTO_LINK.get(reply_id)

    if not client_id:
        bot.send_message(OWNER_ID, "❌ Не удалось найти клиента по этому сообщению.")
        return

    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add("Устраивает", "Не устраивает")

    bot.send_message(
        client_id,
        f"💰 Примерная стоимость пошива: {message.text}\nЦена вас устраивает?",
        reply_markup=markup
    )
    STATE[client_id] = 'AWAIT_PRICE_CONFIRM'
    del PHOTO_LINK[reply_id]

@bot.message_handler(func=lambda m: STATE.get(m.chat.id) == 'AWAIT_PRICE_CONFIRM')
def price_confirm(message):
    if message.text.lower() == "устраивает":
        bot.send_message(message.chat.id, "📄 Пришлите, пожалуйста, реквизиты вашей компании для выставления счёта.", reply_markup=types.ReplyKeyboardRemove())
        STATE[message.chat.id] = 'AWAIT_REQUISITES'
    else:
        bot.send_message(message.chat.id, "Спасибо за интерес! Если что — будем на связи.", reply_markup=types.ReplyKeyboardRemove())
        STATE.pop(message.chat.id, None)
        DATA.pop(message.chat.id, None)

@bot.message_handler(func=lambda m: STATE.get(m.chat.id) == 'AWAIT_REQUISITES')
def handle_requisites(message):
    user = message.from_user
    text = message.text
    photo_path = DATA.get(message.chat.id, {}).get("photo_path")

    if not photo_path:
        bot.send_message(message.chat.id, "❌ Ошибка. Начните сначала: /start")
        return

    save_to_excel(user, photo_path, text)

    caption = (
        f"📬 Клиент прислал реквизиты:\n"
        f"👤 {user.first_name} (@{user.username or '—'})\n"
        f"🆔 {user.id}\n\n"
        f"📄 Реквизиты:\n{text}"
    )

    try:
        with open(photo_path, 'rb') as photo:
            bot.send_photo(OWNER_ID, photo, caption=caption)
    except Exception as e:
        print(f"❌ Ошибка отправки реквизитов владельцу: {e}")

    bot.send_message(message.chat.id, "✅ Спасибо! Мы скоро свяжемся с вами по Telegram.")

    inline = types.InlineKeyboardMarkup()
    inline.add(types.InlineKeyboardButton("📲 Связаться с менеджером", url="https://t.me/xatyba"))
    bot.send_message(message.chat.id, "💬 Если хотите уточнить детали — нажмите кнопку ниже:", reply_markup=inline)

    STATE.pop(message.chat.id, None)
    DATA.pop(message.chat.id, None)

# === Flask + Webhook ===
app = flask.Flask(__name__)

@app.route("/", methods=["GET"])
def index():
    return "Factory Morozov bot is running!", 200

@app.route("/", methods=["POST"])
def webhook():
    if flask.request.headers.get("content-type") == "application/json":
        json_string = flask.request.get_data().decode("utf-8")
        print(f"[WEBHOOK] Получено: {json_string}")
        try:
            update = telebot.types.Update.de_json(json_string)
            bot.process_new_updates([update])
            print("[WEBHOOK] Обработка обновления завершена.")
        except Exception as e:
            print(f"[WEBHOOK] Ошибка при обработке: {e}")
        return "ok", 200
    else:
        return "Unsupported Media Type", 415



bot.remove_webhook()
bot.set_webhook(url="https://factory-morozov-bot.onrender.com")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
